import os
from openpyxl import Workbook  # Import openpyxl for Excel file creation
from datetime import datetime  # Import for date parsing
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from tqdm import tqdm  # Import tqdm for the progress bar
import shutil  # Import for folder deletion
from concurrent.futures import ThreadPoolExecutor, as_completed # Import for multithreading

# Dynamically determine the folder path for "Data"
script_dir = os.path.dirname(os.path.abspath(__file__))  # Get the directory of the script
folder_path = os.path.join(script_dir, "Data")  # Path to the "Data" folder

# Ask the user for the output folder name
output_folder_name = str(input("Enter the name of the output folder (will replace duplicates): ") or "Results").strip()
output_folder_path = os.path.join(script_dir, output_folder_name)

# Path to the unfound IDs file
unfound_ids_file_path = os.path.join(script_dir, "Unfound_IDs.txt") # Path to the unfound IDs file

# Check if the folder already exists and delete it if necessary
if os.path.exists(output_folder_path):
    try:
        shutil.rmtree(output_folder_path)
        print(f"Existing folder '{output_folder_path}' deleted.")
    except PermissionError:
        print(f"Error: Unable to delete the existing folder '{output_folder_path}'. Please close any open files and try again.")
        exit(1)

# Create the new output folder
os.makedirs(output_folder_path, exist_ok=True)
print(f"Output folder '{output_folder_path}' created.")

# Input file containing the list of animal ID file paths
id_file_list_path = os.path.join(script_dir, "ID List.txt")  # Path to the file containing ID file paths

# Read the IDs from "ID List.txt" and assign them to ID_list
ID_list = []
try:
    with open(id_file_list_path, 'r', encoding='utf-8') as id_file_list:
        ID_list = [line.strip() for line in id_file_list.readlines()]  # Read and strip each line
except FileNotFoundError:
    print(f"Error: The file '{id_file_list_path}' was not found.")
    exit(1)
except PermissionError:
    print(f"Error: Permission denied for the file '{id_file_list_path}'.")
    exit(1)


valid_file_processed = True  # Flag to track if any valid files were processed
# Check if any IDs were loaded
if not ID_list:
    valid_file_processed = False  # No valid files were processed
    print("Error: No IDs were found in 'ID List.txt'.")
    exit(1)

def search_string_in_files_with_context(folder_path, search_string):
    """
    Search for a given string in all files within a folder and copy lines from two lines above the match
    until one empty line is encountered.

    Args:
        folder_path (str): Path to the folder to search in.
        search_string (str): The string to search for.

    Returns:
        dict: A dictionary where keys are file paths and values are lists of matched lines with context.
    """
    results = {}

    # Walk through all files in the folder
    for root, _, files in os.walk(folder_path):
        for file in files:
            file_path = os.path.join(root, file)
            try:
                # Open and read the file line by line
                with open(file_path, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                    for i, line in enumerate(lines):
                        if search_string in line:
                            # Start copying from two lines above the match
                            start_index = max(0, i - 2)
                            context = []

                            for j in range(start_index, len(lines)):
                                current_line = lines[j].strip()

                                # Exclude lines that contain "File:" (case-insensitive)
                                if not current_line.lower().startswith("file:"):
                                    context.append(current_line)

                                # Check for consecutive empty lines
                                if current_line == "":  # Stop after an empty line
                                    break

                            if file_path not in results:
                                results[file_path] = []
                            results[file_path].append(context)
            except (UnicodeDecodeError, PermissionError):
                # Skip files that cannot be read
                print(f"Error reading file: {file_path}")
                pass

    return results

def process_id(search_string):
    """
    Process a single ID: search for the ID in files, extract context, and save to an Excel file.
    """
    result = search_string_in_files_with_context(folder_path, search_string.strip("'"))
    if result:
        for file, contexts in result.items():
            # Create a new workbook for each file
            file_name = f"{search_string}_{os.path.splitext(os.path.basename(file))[0]}.xlsx"
            output_file_path = os.path.join(output_folder_path, file_name)

            try:
                # Each thread creates its own Workbook object
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Data"

                for context in contexts:
                    for line_index, line in enumerate(context):
                        # Split the line into columns by spaces
                        columns = line.split()

                        # Handle the first two lines as dates
                        if line_index < 2:
                            try:
                                date_formats = ["%m/%d/%y", "%d-%m-%Y"]
                                date = None
                                for fmt in date_formats:
                                    try:
                                        date = datetime.strptime(columns[2], fmt)
                                        break
                                    except ValueError:
                                        continue
                                if date:
                                    columns[2] = date.strftime("%m/%d/%Y")
                            except (IndexError, ValueError):
                                pass
                        else:
                            # Convert all numeric strings to numbers for subsequent lines
                            for i in range(len(columns)):
                                if isinstance(columns[i], str) and columns[i].isdigit():
                                    columns[i] = int(columns[i])
                                else:
                                    try:
                                        columns[i] = float(columns[i])
                                    except ValueError:
                                        pass

                        # Write the processed row (columns) to the Excel sheet
                        sheet.append(columns)

                # Adjust column widths to fit the content
                for column_cells in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column_cells[0].column)
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    sheet.column_dimensions[column_letter].width = adjusted_width

                # Save the workbook
                workbook.save(output_file_path)
            except Exception as e:
                print(f"Error writing to file '{output_file_path}': {e}")
    else:
        print(f"The ID {search_string} was not found in any files.")
        return search_string  # Return unfound ID for tracking

    return None  # Return None if the ID was processed successfully


if __name__ == "__main__":
    if not valid_file_processed:
        print("No valid IDs found in the specified files.")
    elif not ID_list:
        print("No IDs were found in the valid files provided.")
    else:
        unfound_ids = []  # List to track IDs that were not found

        # Use ThreadPoolExecutor for multithreading
        with ThreadPoolExecutor() as executor:
            futures = {executor.submit(process_id, search_string): search_string for search_string in ID_list}

            # Use tqdm to track progress
            with tqdm(total=len(ID_list), desc="Processing IDs", unit="ID") as pbar:
                for future in as_completed(futures):
                    search_string = futures[future]
                    try:
                        result = future.result()
                        if result:  # If the ID was not found
                            unfound_ids.append(result)
                    except Exception as e:
                        print(f"Error processing ID '{search_string}': {e}")
                    finally:
                        pbar.update(1)

# Check if the unfound IDs file already exists and delete it if necessary
if os.path.exists(unfound_ids_file_path):
    try:
        shutil.rmtree(unfound_ids_file_path)
        print(f"Existing file '{unfound_ids_file_path}' deleted.")
    except PermissionError:
        print(f"Error: Unable to delete the existing file '{unfound_ids_file_path}'. Please close any open files and try again.")
        exit(1)
    os.makedirs(unfound_ids_file_path, exist_ok=True)
    print(f"Output file '{unfound_ids_file_path}' created.")

# Save the list of unfound IDs to a .txt file
try:
    with open(unfound_ids_file_path, 'w', encoding='utf-8') as unfound_file:
        if unfound_ids:
            unfound_file.write("The following IDs were not found:\n")
            for unfound_id in unfound_ids:
                unfound_file.write(f"{unfound_id}\n")
            # Write the unfound IDs to the file
            print(f"\nUnfound IDs have been saved to '{unfound_ids_file_path}'.")
        else:
            unfound_file.write("All IDs were found in the files.")
except Exception as e:
    print(f"Error writing to 'Unfound_IDs.txt': {e}")